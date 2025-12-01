#!/usr/bin/env python3
"""
PE_182.py - Test Case
Test Case ID: PE_182
Test Case Description: To verify prepaid ledger calculation for ST = 47 With Consumption = up to 400 kWh and MD = upto 150%

Step 1: Create account with specified parameters
Step 2: Fill daily load data
Step 3: Fill profile instant data
Step 4: Trigger prepaid ledger APIs
"""

import requests
import random
import string
import psycopg2
import logging
from datetime import datetime, timedelta
import os
from typing import List, Dict, Any
from urllib.parse import quote
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ===== CONFIGURATION =====

# Fixed parameters and consumer configuration
SUPPLY_TYPECODE = "47"
SANCTIONED_LOAD = 12
LOAD_UNIT = "KVA"
METER_INSTALL_DATE = "2025-11-01T00:00:00"
NET_METER_FLAG = "N"
GREEN_ENERGY_FLAG = "N"
POWER_LOOM_COUNT = 0
PREPAID_OPENING_BALANCE = 4000
PARAM1 = "0"
PREPAID_POSTPAID_FLAG = "1"
ED_APPLICABLE = "1"

# Daily load and profile instant configuration
END_WH = 400000
MD_W = 15000

# API Configuration
ACCOUNT_API_URL = "https://integration1.stage.gomatimvvnl.in/initial_master_sync/"
ACCOUNT_API_HEADERS = {
    'Content-Type': 'application/json',
    'Accept': 'application/json',
    'Authorization': 'Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJ1c2VybmFtZSI6IlZhbGlkYXRpb250ZWFtIn0.cSMyfD9Cz7LJQnbRIsrPCtOif_ZgbucMHra7BoKXGGg'
}

LEDGER_API_BASE = "https://engine-web.stage.gomatimvvnl.in"
# MDMS API Configuration
MDMS_API_BASE = "https://mdms-api.stage.gomatimvvnl.in/db-service"
DAILYLOAD_API_URL = f"{MDMS_API_BASE}/dailyloads"
PROFILE_INSTANT_API_URL = f"{MDMS_API_BASE}/profileinstant"


# Database Configuration
DB_CONFIG = {
    'host': 'db_mdms.stage.mvvnl.internal',
    'database': 'validation_rules',
    'user': 'postgres',
    'password': 'LBjsKJkd937042!!',
    'port': 5432
}

# Table Names
DAILYLOAD_TABLE = 'dailyload_vee_validated'
PROFILE_INSTANT_TABLE = 'profile_instant_vee'

# Logging Configuration
os.makedirs("logs", exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('logs/PE_182.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ===== HELPER FUNCTIONS =====

def generate_random_number(length):
    """Generate a random number string of specified length"""
    return ''.join([str(random.randint(0, 9)) for _ in range(length)])

def generate_random_alphanumeric(length):
    """Generate a random alphanumeric string of specified length"""
    characters = string.ascii_uppercase + string.digits
    return ''.join(random.choice(characters) for _ in range(length))

def generate_consumer_name():
    """Generate a random consumer name"""
    first_names = ["John", "Jane", "Robert", "Mary", "James", "Patricia", "Michael", "Jennifer", 
                   "William", "Linda", "David", "Elizabeth", "Richard", "Barbara", "Joseph", "Susan"]
    last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis",
                  "Rodriguez", "Martinez", "Hernandez", "Lopez", "Wilson", "Anderson", "Thomas", "Taylor"]
    return f"{random.choice(first_names)} {random.choice(last_names)}"

def generate_email(consumer_name):
    """Generate an email based on consumer name and 3 digit number"""
    email_name = consumer_name.lower().replace(" ", "")
    three_digit_number = generate_random_number(3)
    return f"{email_name}{three_digit_number}@example.com"

def generate_mobile_number():
    """Generate a random 10-digit mobile number"""
    return int(generate_random_number(10))

def parse_install_date(date_str: str) -> str:
    """Parse meterInstalldate from format (2025-10-01T00:00:00) to (yyyy-mm-dd 00:00:00)."""
    try:
        if 'T' in date_str:
            dt = datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%S')
        else:
            dt = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
        return dt.strftime('%Y-%m-%d %H:%M:%S')
    except Exception as e:
        logger.error(f"Error parsing date {date_str}: {e}")
        return None

def get_month_end_date(start_date_str: str) -> str:
    """Calculate the first day of the next month from the given start date (month end + 1 day)."""
    try:
        dt = datetime.strptime(start_date_str, '%Y-%m-%d %H:%M:%S')
        if dt.month == 12:
            next_month = 1
            next_year = dt.year + 1
        else:
            next_month = dt.month + 1
            next_year = dt.year
        month_end = datetime(next_year, next_month, 1, 0, 0, 0)
        return month_end.strftime('%Y-%m-%d %H:%M:%S')
    except Exception as e:
        logger.error(f"Error calculating month end date from {start_date_str}: {e}")
        return None

# ===== STEP 1: CREATE ACCOUNT =====

def create_account():
    """Step 1: Create account with specified parameters"""
    logger.info("=" * 60)
    logger.info("STEP 1: Creating Account")
    logger.info("=" * 60)
    
    # Generate accountId (10 digits)
    account_id = generate_random_number(10)
    logger.info(f"Generated Account ID: {account_id}")
    
    # Generate meterSrno = PO + (last 8 digits of account id)
    last_8_digits = account_id[-8:]
    meter_srno = f"PO{last_8_digits}"
    logger.info(f"Generated Meter Serial No: {meter_srno}")
    
    # Generate additional required fields
    request_id = generate_random_number(8)
    consumer_name = generate_consumer_name()
    mobile_number = generate_mobile_number()
    email = generate_email(consumer_name)
    badge_number = f"GPMPO{last_8_digits}"
    
    logger.info(f"Using configuration parameters:")
    logger.info(f"  Supply Type Code: {SUPPLY_TYPECODE}")
    logger.info(f"  Sanctioned Load: {SANCTIONED_LOAD}")
    logger.info(f"  Load Unit: {LOAD_UNIT}")
    logger.info(f"  Meter Install Date: {METER_INSTALL_DATE}")
    logger.info(f"  Prepaid Opening Balance: {PREPAID_OPENING_BALANCE}")
    
    # Construct payload
    payload = {
        "requestId": request_id,
        "accountId": account_id,
        "consumerName": consumer_name,
        "address1": "Lucknow",
        "address2": "UP",
        "address3": "India",
        "postcode": int(generate_random_number(6)),
        "mobileNumber": int(mobile_number),
        "whatsAppnumber": int(mobile_number),
        "email": email,
        "badgeNumber": badge_number,
        "supplyTypecode": SUPPLY_TYPECODE,
        "meterSrno": meter_srno,
        "sanctionedLoad": float(SANCTIONED_LOAD),
        "loadUnit": LOAD_UNIT,
        "meterInstalldate": METER_INSTALL_DATE,
        "customerEntrydate": METER_INSTALL_DATE.split("T")[0],
        "connectionStatus": "C",
        "prepaidPostpaidflag": PREPAID_POSTPAID_FLAG,
        "netMeterflag": NET_METER_FLAG,
        "shuntCapacitorflag": "N",
        "greenEnergyflag": GREEN_ENERGY_FLAG,
        "powerLoomcount": int(POWER_LOOM_COUNT),
        "rateSchedule": generate_random_alphanumeric(6),
        "meterType": "1PH-STSM",
        "meterMake": "Polaris",
        "multiplyingFactor": float(1),
        "meterStatus": "A",
        "arrears": float(0.00),
        "prepaidOpeningbalance": float(PREPAID_OPENING_BALANCE),
        "divisionCode": f"DIV{generate_random_alphanumeric(7)}",
        "subDivCode": f"SUBDIV{generate_random_alphanumeric(4)}",
        "dtrCode": "Engine",
        "feederCode": "Engine",
        "substaionCode": "Engine",
        "serviceAgreementid": "",
        "billCycle": "Monthly",
        "edApplicable": ED_APPLICABLE,
        "lpsc": float(0.00),
        "param1": PARAM1,
        "param2": "01-04-2024",
        "param3": "0",
        "param4": "string",
        "param5": "string",
        "consumerType": "P",
        "fromDate": "01-04-2024",
        "toDate": "01-04-2024",
        "offSeasonstartDate": "01-04-2024",
        "offSeasonendDate": "01-04-2024",
        "offSeasonbenifitFlag": "N",
        "offSeasonviolationCount": "0",
        "offSeasonload": float(0.00)
    }
    
    try:
        logger.info(f"Creating account with Account ID: {account_id}, Meter Serial No: {meter_srno}")
        response = requests.post(ACCOUNT_API_URL, headers=ACCOUNT_API_HEADERS, json=payload)
        logger.info(f"Account creation response: {response.status_code}")
        logger.info(f"Response text: {response.text}")
        
        if response.status_code == 200:
            logger.info("Account created successfully!")
            return {
                'accountId': account_id,
                'meterSrno': meter_srno,
                'meterInstalldate': METER_INSTALL_DATE,
                'badgeNumber': badge_number,
                'payload': payload
            }
        else:
            logger.error(f"Account creation failed with status {response.status_code}")
            return None
    except Exception as e:
        logger.error(f"Error creating account: {str(e)}")
        return None

# ===== STEP 2: FILL DAILY LOAD DATA =====

def fill_daily_load_data(account_data):
    """Step 2: Fill daily load data"""
    logger.info("=" * 60)
    logger.info("STEP 2: Filling Daily Load Data")
    logger.info("=" * 60)
    
    meter_srno = account_data['meterSrno']
    meter_install_date = account_data['meterInstalldate']
    
    # Parse dates
    start_date_str = parse_install_date(meter_install_date)
    end_date_str = get_month_end_date(start_date_str)
    
    logger.info(f"Device Identifier: {meter_srno}")
    logger.info(f"Start Date: {start_date_str}")
    logger.info(f"End Date: {end_date_str}")
    logger.info(f"End Wh: {END_WH}")
    
    # Connect to database
    try:
        connection = psycopg2.connect(**DB_CONFIG)
        cursor = connection.cursor()
        logger.info("Connected to database")
        
        # Parse dates for calculation
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d %H:%M:%S') - timedelta(hours=5, minutes=30)
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d %H:%M:%S') - timedelta(hours=5, minutes=30)
        total_days = (end_date - start_date).days + 1
        
        logger.info(f"Total days to process: {total_days}")
        
        # Generate data
        data_list = []
        current_wh = 0
        current_vah = 0
        current_export_wh = 0
        current_export_vah = 0
        end_vah = END_WH / 0.90
        end_export_wh = 750
        end_export_vah = end_export_wh / 0.90
        
        logger.info(f"Starting data generation with end values - Wh: {END_WH}, VAh: {end_vah:.2f}")
        
        for day in range(total_days):
            current_date = start_date + timedelta(days=day)
            remaining_days = total_days - day - 1
            
            if day == total_days - 1:
                current_wh = int(END_WH)
                current_vah = int(end_vah)
                current_export_wh = int(end_export_wh)
                current_export_vah = int(end_export_vah)
            else:
                wh_needed = END_WH - current_wh
                vah_needed = end_vah - current_vah
                export_wh_needed = end_export_wh - current_export_wh
                export_vah_needed = end_export_vah - current_export_vah
                
                avg_increment_wh = wh_needed / remaining_days if remaining_days > 0 else 0
                avg_increment_vah = vah_needed / remaining_days if remaining_days > 0 else 0
                avg_increment_export_wh = export_wh_needed / remaining_days if remaining_days > 0 else 0
                avg_increment_export_vah = export_vah_needed / remaining_days if remaining_days > 0 else 0
                
                min_factor = 0.3
                max_factor = 1.7
                
                min_increment_wh = max(0, int(avg_increment_wh * min_factor))
                max_increment_wh = min(int(avg_increment_wh * max_factor), int(wh_needed))
                increment_wh = random.randint(min_increment_wh, max_increment_wh) if max_increment_wh > min_increment_wh else min_increment_wh
                current_wh += increment_wh
                
                min_increment_vah = max(0, int(avg_increment_vah * min_factor))
                max_increment_vah = min(int(avg_increment_vah * max_factor), int(vah_needed))
                increment_vah = random.randint(min_increment_vah, max_increment_vah) if max_increment_vah > min_increment_vah else min_increment_vah
                current_vah += increment_vah
                
                min_increment_export_wh = max(0, int(avg_increment_export_wh * min_factor))
                max_increment_export_wh = min(int(avg_increment_export_wh * max_factor), int(export_wh_needed))
                increment_export_wh = random.randint(min_increment_export_wh, max_increment_export_wh) if max_increment_export_wh > min_increment_export_wh else min_increment_export_wh
                current_export_wh += increment_export_wh
                
                min_increment_export_vah = max(0, int(avg_increment_export_vah * min_factor))
                max_increment_export_vah = min(int(avg_increment_export_vah * max_factor), int(export_vah_needed))
                increment_export_vah = random.randint(min_increment_export_vah, max_increment_export_vah) if max_increment_export_vah > min_increment_export_vah else min_increment_export_vah
                current_export_vah += increment_export_vah
            
            data = {
                'device_id': None,
                'data_timestamp': current_date.strftime('%Y-%m-%d %H:%M:%S'),
                'dcu_serial': f'DCU{meter_srno[-6:]}',
                'device_identifier': meter_srno,
                'data_source': 'PUSH',
                'data_type': 'DLFV',
                'import_Wh': current_wh,
                'import_VAh': current_vah,
                'export_Wh': current_export_wh,
                'export_VAh': current_export_vah,
                'exec_datetime': current_date.strftime('%Y-%m-%d %H:%M:%S'),
                'dailyload_datetime': current_date.strftime('%Y-%m-%d %H:%M:%S'),
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'is_active': True,
                'is_valid': True,
                'is_estimated': False,
                'is_edited': False,
                'meter_type': '1-ph'
            }
            data_list.append(data)
            
            # Log progress every 10 days or on first/last day
            if day == 0 or day == total_days - 1 or day % 10 == 0:
                logger.info(f"Day {day + 1}/{total_days}: {current_date.strftime('%Y-%m-%d')} - Wh: {current_wh}, VAh: {current_vah:.2f}")
        
        logger.info(f"Generated {len(data_list)} daily load records")
        
        # Insert data
        insert_query = f"""
            INSERT INTO {DAILYLOAD_TABLE} (
                device_id, data_timestamp, dcu_serial, device_identifier,
                data_source, data_type, "import_Wh", "import_VAh",
                "export_Wh", "export_VAh", exec_datetime, dailyload_datetime,
                created_at, updated_at, is_active, is_valid,
                is_estimated, is_edited, meter_type
            ) VALUES (
                %(device_id)s, %(data_timestamp)s, %(dcu_serial)s, %(device_identifier)s,
                %(data_source)s, %(data_type)s, %(import_Wh)s, %(import_VAh)s,
                %(export_Wh)s, %(export_VAh)s, %(exec_datetime)s, %(dailyload_datetime)s,
                %(created_at)s, %(updated_at)s, %(is_active)s, %(is_valid)s,
                %(is_estimated)s, %(is_edited)s, %(meter_type)s
            )
        """
        
        cursor.executemany(insert_query, data_list)
        connection.commit()
        
        logger.info(f"Successfully inserted {len(data_list)} daily load records into {DAILYLOAD_TABLE}")
        cursor.close()
        connection.close()
        return True
        
    except Exception as e:
        logger.error(f"Error filling daily load data: {str(e)}")
        if 'connection' in locals():
            connection.rollback()
            connection.close()
        return False

# ===== STEP 3: FILL PROFILE INSTANT DATA =====

def fill_profile_instant_data(account_data):
    """Step 3: Fill profile instant data"""
    logger.info("=" * 60)
    logger.info("STEP 3: Filling Profile Instant Data")
    logger.info("=" * 60)
    
    meter_srno = account_data['meterSrno']
    meter_install_date = account_data['meterInstalldate']
    
    # Parse dates
    start_date_str = parse_install_date(meter_install_date)
    end_date_str = get_month_end_date(start_date_str)
    
    logger.info(f"Device Identifier: {meter_srno}")
    logger.info(f"Start Date: {start_date_str}")
    logger.info(f"End Date: {end_date_str}")
    logger.info(f"MD_W: {MD_W}")
    
    # Connect to database
    try:
        connection = psycopg2.connect(**DB_CONFIG)
        cursor = connection.cursor()
        logger.info("Connected to database")
        
        # Parse dates for calculation
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d %H:%M:%S') - timedelta(hours=5, minutes=30)
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d %H:%M:%S') - timedelta(hours=5, minutes=30)
        total_days = (end_date - start_date).days + 1
        
        logger.info(f"Total days to process: {total_days}")
        
        # Generate data
        data_list = []
        md_va = MD_W / 0.90
        
        logger.info(f"Starting data generation with MD_W: {MD_W}, MD_VA: {md_va:.2f}")
        
        for day in range(total_days):
            current_date = start_date + timedelta(days=day)
            progress_ratio = day / (total_days - 1) if total_days > 1 else 1
            current_md_w = int(MD_W * progress_ratio)
            current_md_va = int(md_va * progress_ratio)
            
            data = {
                'meter_id': 0,
                'device_identifier': meter_srno,
                'data_timestamp': current_date.strftime('%Y-%m-%d %H:%M:%S'),
                'create_timestamp': current_date.strftime('%Y-%m-%d %H:%M:%S'),
                'exec_datetime': current_date.strftime('%Y-%m-%d %H:%M:%S'),
                'meter_type': '1-ph',
                'import_Wh': 0,
                'export_Wh': 0,
                'import_VAh': 0,
                'export_VAh': 0,
                'cumm_energy_VArh_Q1': 0,
                'cumm_energy_VArh_Q2': 0,
                'cumm_energy_VArh_Q3': 0,
                'cumm_energy_VArh_Q4': 0,
                'signed_active_power_KVAR': 0,
                'MD_W': current_md_w,
                'MD_VA': current_md_va,
                'MD_W_datetime': current_date.strftime('%Y-%m-%d %H:%M:%S'),
                'MD_VA_datetime': current_date.strftime('%Y-%m-%d %H:%M:%S'),
                'PF': 1,
                'Rphase_signed_PF': 0,
                'Yphase_signed_PF': 0,
                'Bphase_signed_PF': 0,
                'voltage': 230,
                'RN_voltage': 0,
                'YN_voltage': 0,
                'BN_voltage': 0,
                'phase_current': 0,
                'Rphase_current': 0,
                'Yphase_current': 0,
                'Bphase_current': 0,
                'neutral_current': 0,
                'frequency': 50.0,
                'active_power_W': 0,
                'apparent_power_VA': 0,
                'load_limit_value': 0,
                'load_limit_func_status': False,
                'cumm_tamper_count': 0,
                'cumm_billing_count': 0,
                'cumm_programming_count': 0,
                'num_power_fail': 0,
                'num_power_fail_dur': 0,
                'meter_current_datetime': current_date.strftime('%Y-%m-%d %H:%M:%S'),
                'last_billing_datetime': current_date.strftime('%Y-%m-%d %H:%M:%S'),
                'cumm_power_on_dur_minute': 0,
                'is_valid': True,
                'is_estimated': False,
                'is_edited': False,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'device_id': None,
                'file_log_id': 0,
                'dcu_serial': f'DCU{meter_srno[-6:]}',
                'hes_pk': '0',
                'hes_msg_id': None,
                'hes_created_at': 0,
                'hes_created_at_str': current_date.strftime('%Y-%m-%d %H:%M:%S')
            }
            data_list.append(data)
            
            # Log progress every 10 days or on first/last day
            if day == 0 or day == total_days - 1 or day % 10 == 0:
                logger.info(f"Day {day + 1}/{total_days}: {current_date.strftime('%Y-%m-%d')} - MD_W: {current_md_w}, MD_VA: {current_md_va}")
        
        logger.info(f"Generated {len(data_list)} profile instant records")
        
        # Insert data
        insert_query = f"""
            INSERT INTO {PROFILE_INSTANT_TABLE} (
                meter_id, device_identifier, data_timestamp, create_timestamp, exec_datetime, meter_type,
                "import_Wh", "export_Wh", "import_VAh", "export_VAh",
                "cumm_energy_VArh_Q1", "cumm_energy_VArh_Q2", "cumm_energy_VArh_Q3", "cumm_energy_VArh_Q4",
                "signed_active_power_KVAR", "MD_W", "MD_VA", "MD_W_datetime", "MD_VA_datetime",
                "PF", "Rphase_signed_PF", "Yphase_signed_PF", "Bphase_signed_PF",
                voltage, "RN_voltage", "YN_voltage", "BN_voltage",
                phase_current, "Rphase_current", "Yphase_current", "Bphase_current",
                neutral_current, frequency, "active_power_W", "apparent_power_VA",
                load_limit_value, load_limit_func_status,
                cumm_tamper_count, cumm_billing_count, cumm_programming_count,
                num_power_fail, num_power_fail_dur,
                meter_current_datetime, last_billing_datetime, cumm_power_on_dur_minute,
                is_valid, is_estimated, is_edited,
                created_at, updated_at,
                device_id, file_log_id, dcu_serial,
                hes_pk, hes_msg_id, hes_created_at
            ) VALUES (
                %(meter_id)s, %(device_identifier)s, %(data_timestamp)s, %(create_timestamp)s, %(exec_datetime)s, %(meter_type)s,
                %(import_Wh)s, %(export_Wh)s, %(import_VAh)s, %(export_VAh)s,
                %(cumm_energy_VArh_Q1)s, %(cumm_energy_VArh_Q2)s, %(cumm_energy_VArh_Q3)s, %(cumm_energy_VArh_Q4)s,
                %(signed_active_power_KVAR)s, %(MD_W)s, %(MD_VA)s, %(MD_W_datetime)s, %(MD_VA_datetime)s,
                %(PF)s, %(Rphase_signed_PF)s, %(Yphase_signed_PF)s, %(Bphase_signed_PF)s,
                %(voltage)s, %(RN_voltage)s, %(YN_voltage)s, %(BN_voltage)s,
                %(phase_current)s, %(Rphase_current)s, %(Yphase_current)s, %(Bphase_current)s,
                %(neutral_current)s, %(frequency)s, %(active_power_W)s, %(apparent_power_VA)s,
                %(load_limit_value)s, %(load_limit_func_status)s,
                %(cumm_tamper_count)s, %(cumm_billing_count)s, %(cumm_programming_count)s,
                %(num_power_fail)s, %(num_power_fail_dur)s,
                %(meter_current_datetime)s, %(last_billing_datetime)s, %(cumm_power_on_dur_minute)s,
                %(is_valid)s, %(is_estimated)s, %(is_edited)s,
                %(created_at)s, %(updated_at)s,
                %(device_id)s, %(file_log_id)s, %(dcu_serial)s,
                %(hes_pk)s, %(hes_msg_id)s, %(hes_created_at_str)s
            )
        """
        
        cursor.executemany(insert_query, data_list)
        connection.commit()
        
        logger.info(f"Successfully inserted {len(data_list)} profile instant records into {PROFILE_INSTANT_TABLE}")
        cursor.close()
        connection.close()
        return True
        
    except Exception as e:
        logger.error(f"Error filling profile instant data: {str(e)}")
        if 'connection' in locals():
            connection.rollback()
            connection.close()
        return False

# ===== STEP 4: TRIGGER PREPAID LEDGER APIs =====

def trigger_prepaid_ledger(account_id):
    """Step 4: Trigger prepaid ledger APIs"""
    logger.info("=" * 60)
    logger.info("STEP 4: Triggering Prepaid Ledger APIs")
    logger.info("=" * 60)
    
    # Get yesterday's date in YYYY-MM-DD HH:MM:SS format (with 00:00:00 for time)
    yesterday = datetime.now() - timedelta(days=1)
    date_str = yesterday.strftime('%Y-%m-%d 00:00:00')
    # URL encode the date for the API
    date_encoded = quote(date_str)
    
    logger.info(f"Account ID: {account_id}")
    logger.info(f"Date (yesterday): {date_str}")
    logger.info(f"URL encoded date: {date_encoded}")
    
    # First API: daily_ledger_task
    try:
        url1 = f"{LEDGER_API_BASE}/trigger_task/daily_ledger_task/{date_encoded}?wallet_balance_sync_flag=False&account_id={account_id}"
        logger.info(f"Calling API 1: {url1}")
        
        response1 = requests.get(
            url1,
            headers={'accept': 'application/json'}
        )
        
        logger.info(f"API 1 Response Status: {response1.status_code}")
        logger.info(f"API 1 Response Text: {response1.text}")
        
        if response1.status_code == 200:
            logger.info("API 1 (daily_ledger_task) called successfully")
            return True
        else:
            logger.warning(f"API 1 returned status {response1.status_code}")
            return False
    except Exception as e:
        logger.error(f"Error calling API 1: {str(e)}")
        return False

# ===== STEP 5: GENERATE EXCEL REPORT =====

def fetch_daily_load_data(meter_srno):
    '''Fetch daily load data from MDMS API'''
    try:
        payload = {
            "badge_numbers": [meter_srno],
            "page": 1,
            "limit": 100
        }
        response = requests.post(DAILYLOAD_API_URL, json=payload, headers={'accept': 'application/json', 'Content-Type': 'application/json'})
        
        if response.status_code == 200:
            data = response.json()
            if 'data' in data and data['data']:
                return data['data']
        logger.warning(f"Failed to fetch daily load data: {response.status_code}")
        return []
    except Exception as e:
        logger.error(f"Error fetching daily load data: {str(e)}")
        return []

def fetch_profile_instant_data(meter_srno):
    '''Fetch profile instant data from MDMS API'''
    try:
        payload = {
            "badge_numbers": [meter_srno],
            "page": 1,
            "limit": 100
        }
        response = requests.post(PROFILE_INSTANT_API_URL, json=payload, headers={'accept': 'application/json', 'Content-Type': 'application/json'})
        
        if response.status_code == 200:
            data = response.json()
            if 'data' in data and data['data']:
                return data['data']
        logger.warning(f"Failed to fetch profile instant data: {response.status_code}")
        return []
    except Exception as e:
        logger.error(f"Error fetching profile instant data: {str(e)}")
        return []

def generate_excel_report(account_data):
    '''Step 5: Generate Excel report with all test data'''
    logger.info("=" * 60)
    logger.info("STEP 5: Generating Excel Report")
    logger.info("=" * 60)
    
    # Create Result_File folder
    result_folder = "Result_File"
    os.makedirs(result_folder, exist_ok=True)
    logger.info(f"Created/Verified folder: {result_folder}")
    
    # Excel file path - will be set dynamically
    excel_file = os.path.join(result_folder, "Report_PE_182.xlsx")
    
    try:
        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # ===== SHEET 1: SUMMARY =====
        ws_summary = wb.create_sheet("Summary", 0)
        
        # Test case information - will be extracted from file
        test_case_id = "PE_182"
        test_case_description = "To verify prepaid ledger calculation for ST = 47 With Consumption = up to 400 kWh and MD = upto 150%"
        account_id = account_data['accountId']
        meter_number = account_data['meterSrno']
        badge_number = account_data['badgeNumber']
        status = "Completed" if account_data else "Failed"
        
        # Headers
        headers = ["Test Case ID", "Test Case Description", "Account ID", "Meter Number", "Badge Number", "Status"]
        ws_summary.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data row
        data_row = [test_case_id, test_case_description, account_id, meter_number, badge_number, status]
        ws_summary.append(data_row)
        
        # Auto-adjust column widths
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Summary sheet created")
        
        # ===== SHEET 2: CONSUMER DETAILS =====
        ws_consumer = wb.create_sheet("Consumer Details", 1)
        
        payload = account_data['payload']
        
        # Write headers first
        ws_consumer.append(["Parameter", "Value"])
        
        # Convert payload to rows (key-value pairs) and write data
        for key, value in payload.items():
            ws_consumer.append([key, str(value)])
        
        # Style headers
        for cell in ws_consumer[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in ws_consumer.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_consumer.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Consumer Details sheet created")
        
        # ===== SHEET 3: DAILY LOAD =====
        ws_daily = wb.create_sheet("Daily Load", 2)
        
        # Fetch daily load data
        logger.info(f"Fetching daily load data for meter: {meter_number}")
        daily_load_data = fetch_daily_load_data(meter_number)
        
        if daily_load_data:
            # Extract only required fields
            daily_load_filtered = []
            for record in daily_load_data:
                filtered_record = {
                    "dailyload_datetime": record.get("dailyload_datetime", ""),
                    "badge_number": record.get("badge_number", ""),
                    "meter_serial_number": record.get("meter_serial_number", ""),
                    "data_source": record.get("data_source", ""),
                    "data_type": record.get("data_type", ""),
                    "export_Wh": record.get("export_Wh", ""),
                    "import_Wh": record.get("import_Wh", ""),
                    "export_VAh": record.get("export_VAh", ""),
                    "import_VAh": record.get("import_VAh", "")
                }
                daily_load_filtered.append(filtered_record)
            
            df_daily = pd.DataFrame(daily_load_filtered)
            
            # Write headers first
            headers = list(df_daily.columns)
            ws_daily.append(headers)
            
            # Write data rows
            for row in df_daily.itertuples(index=False):
                ws_daily.append(list(row))
            
            # Style headers
            for cell in ws_daily[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            logger.info(f"Added {len(daily_load_filtered)} daily load records")
        else:
            ws_daily.append(["No data available"])
            logger.warning("No daily load data found")
        
        # Auto-adjust column widths
        for column in ws_daily.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_daily.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Daily Load sheet created")
        
        # ===== SHEET 4: PROFILE INSTANT =====
        ws_profile = wb.create_sheet("Profile Instant", 3)
        
        # Fetch profile instant data
        logger.info(f"Fetching profile instant data for meter: {meter_number}")
        profile_instant_data = fetch_profile_instant_data(meter_number)
        
        if profile_instant_data:
            # Extract only required fields
            profile_filtered = []
            for record in profile_instant_data:
                filtered_record = {
                    "data_timestamp": record.get("data_timestamp", ""),
                    "badge_number": record.get("badge_number", ""),
                    "meter_serial_number": record.get("meter_serial_number", ""),
                    "meter_type": record.get("meter_type", ""),
                    "MD_W": record.get("MD_W", ""),
                    "MD_VA": record.get("MD_VA", ""),
                    "MD_W_datetime": record.get("MD_W_datetime", ""),
                    "MD_VA_datetime": record.get("MD_VA_datetime", ""),
                    "voltage": record.get("voltage", ""),
                    "frequency": record.get("frequency", "")
                }
                profile_filtered.append(filtered_record)
            
            df_profile = pd.DataFrame(profile_filtered)
            
            # Write headers first
            headers = list(df_profile.columns)
            ws_profile.append(headers)
            
            # Write data rows
            for row in df_profile.itertuples(index=False):
                ws_profile.append(list(row))
            
            # Style headers
            for cell in ws_profile[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            logger.info(f"Added {len(profile_filtered)} profile instant records")
        else:
            ws_profile.append(["No data available"])
            logger.warning("No profile instant data found")
        
        # Auto-adjust column widths
        for column in ws_profile.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_profile.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Profile Instant sheet created")
        
        # Save workbook
        wb.save(excel_file)
        logger.info(f"Excel report saved successfully: {excel_file}")
        return True
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")
        return False


# ===== STEP 5: GENERATE EXCEL REPORT =====

def fetch_daily_load_data(meter_srno):
    '''Fetch daily load data from MDMS API'''
    try:
        payload = {
            "badge_numbers": [meter_srno],
            "page": 1,
            "limit": 100
        }
        response = requests.post(DAILYLOAD_API_URL, json=payload, headers={'accept': 'application/json', 'Content-Type': 'application/json'})
        
        if response.status_code == 200:
            data = response.json()
            if 'data' in data and data['data']:
                return data['data']
        logger.warning(f"Failed to fetch daily load data: {response.status_code}")
        return []
    except Exception as e:
        logger.error(f"Error fetching daily load data: {str(e)}")
        return []

def fetch_profile_instant_data(meter_srno):
    '''Fetch profile instant data from MDMS API'''
    try:
        payload = {
            "badge_numbers": [meter_srno],
            "page": 1,
            "limit": 100
        }
        response = requests.post(PROFILE_INSTANT_API_URL, json=payload, headers={'accept': 'application/json', 'Content-Type': 'application/json'})
        
        if response.status_code == 200:
            data = response.json()
            if 'data' in data and data['data']:
                return data['data']
        logger.warning(f"Failed to fetch profile instant data: {response.status_code}")
        return []
    except Exception as e:
        logger.error(f"Error fetching profile instant data: {str(e)}")
        return []

def generate_excel_report(account_data):
    '''Step 5: Generate Excel report with all test data'''
    logger.info("=" * 60)
    logger.info("STEP 5: Generating Excel Report")
    logger.info("=" * 60)
    
    # Create Result_File folder
    result_folder = "Result_File"
    os.makedirs(result_folder, exist_ok=True)
    logger.info(f"Created/Verified folder: {result_folder}")
    
    # Excel file path - will be set dynamically
    excel_file = os.path.join(result_folder, "Report_PE_182.xlsx")
    
    try:
        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # ===== SHEET 1: SUMMARY =====
        ws_summary = wb.create_sheet("Summary", 0)
        
        # Test case information - will be extracted from file
        test_case_id = "PE_182"
        test_case_description = "To verify prepaid ledger calculation for ST = 47 With Consumption = up to 400 kWh and MD = upto 150%"
        account_id = account_data['accountId']
        meter_number = account_data['meterSrno']
        badge_number = account_data['badgeNumber']
        status = "Completed" if account_data else "Failed"
        
        # Headers
        headers = ["Test Case ID", "Test Case Description", "Account ID", "Meter Number", "Badge Number", "Status"]
        ws_summary.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Data row
        data_row = [test_case_id, test_case_description, account_id, meter_number, badge_number, status]
        ws_summary.append(data_row)
        
        # Auto-adjust column widths
        for column in ws_summary.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_summary.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Summary sheet created")
        
        # ===== SHEET 2: CONSUMER DETAILS =====
        ws_consumer = wb.create_sheet("Consumer Details", 1)
        
        payload = account_data['payload']
        
        # Write headers first
        ws_consumer.append(["Parameter", "Value"])
        
        # Convert payload to rows (key-value pairs) and write data
        for key, value in payload.items():
            ws_consumer.append([key, str(value)])
        
        # Style headers
        for cell in ws_consumer[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in ws_consumer.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_consumer.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Consumer Details sheet created")
        
        # ===== SHEET 3: DAILY LOAD =====
        ws_daily = wb.create_sheet("Daily Load", 2)
        
        # Fetch daily load data
        logger.info(f"Fetching daily load data for meter: {meter_number}")
        daily_load_data = fetch_daily_load_data(meter_number)
        
        if daily_load_data:
            # Extract only required fields
            daily_load_filtered = []
            for record in daily_load_data:
                filtered_record = {
                    "dailyload_datetime": record.get("dailyload_datetime", ""),
                    "badge_number": record.get("badge_number", ""),
                    "meter_serial_number": record.get("meter_serial_number", ""),
                    "data_source": record.get("data_source", ""),
                    "data_type": record.get("data_type", ""),
                    "export_Wh": record.get("export_Wh", ""),
                    "import_Wh": record.get("import_Wh", ""),
                    "export_VAh": record.get("export_VAh", ""),
                    "import_VAh": record.get("import_VAh", "")
                }
                daily_load_filtered.append(filtered_record)
            
            df_daily = pd.DataFrame(daily_load_filtered)
            
            # Write headers first
            headers = list(df_daily.columns)
            ws_daily.append(headers)
            
            # Write data rows
            for row in df_daily.itertuples(index=False):
                ws_daily.append(list(row))
            
            # Style headers
            for cell in ws_daily[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            logger.info(f"Added {len(daily_load_filtered)} daily load records")
        else:
            ws_daily.append(["No data available"])
            logger.warning("No daily load data found")
        
        # Auto-adjust column widths
        for column in ws_daily.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_daily.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Daily Load sheet created")
        
        # ===== SHEET 4: PROFILE INSTANT =====
        ws_profile = wb.create_sheet("Profile Instant", 3)
        
        # Fetch profile instant data
        logger.info(f"Fetching profile instant data for meter: {meter_number}")
        profile_instant_data = fetch_profile_instant_data(meter_number)
        
        if profile_instant_data:
            # Extract only required fields
            profile_filtered = []
            for record in profile_instant_data:
                filtered_record = {
                    "data_timestamp": record.get("data_timestamp", ""),
                    "badge_number": record.get("badge_number", ""),
                    "meter_serial_number": record.get("meter_serial_number", ""),
                    "meter_type": record.get("meter_type", ""),
                    "MD_W": record.get("MD_W", ""),
                    "MD_VA": record.get("MD_VA", ""),
                    "MD_W_datetime": record.get("MD_W_datetime", ""),
                    "MD_VA_datetime": record.get("MD_VA_datetime", ""),
                    "voltage": record.get("voltage", ""),
                    "frequency": record.get("frequency", "")
                }
                profile_filtered.append(filtered_record)
            
            df_profile = pd.DataFrame(profile_filtered)
            
            # Write headers first
            headers = list(df_profile.columns)
            ws_profile.append(headers)
            
            # Write data rows
            for row in df_profile.itertuples(index=False):
                ws_profile.append(list(row))
            
            # Style headers
            for cell in ws_profile[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            logger.info(f"Added {len(profile_filtered)} profile instant records")
        else:
            ws_profile.append(["No data available"])
            logger.warning("No profile instant data found")
        
        # Auto-adjust column widths
        for column in ws_profile.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_profile.column_dimensions[column_letter].width = adjusted_width
        
        logger.info("Profile Instant sheet created")
        
        # Save workbook
        wb.save(excel_file)
        logger.info(f"Excel report saved successfully: {excel_file}")
        return True
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {str(e)}")
        return False

# ===== MAIN FUNCTION =====

def main():
    """Main function to execute all steps"""
    logger.info("=" * 80)
    logger.info("PE_182 Test Plan - Starting Complete Process")
    logger.info("=" * 80)
    
    # Step 1: Create account
    logger.info("Initiating Step 1: Account Creation")
    account_data = create_account()
    if not account_data:
        logger.error("Step 1 failed: Account creation failed. Aborting.")
        return
    
    logger.info("Step 1 completed successfully")
    logger.info(f"Account ID: {account_data['accountId']}")
    logger.info(f"Meter Serial No: {account_data['meterSrno']}")
    
    # Step 2: Fill daily load data
    logger.info("Initiating Step 2: Daily Load Data Filling")
    if not fill_daily_load_data(account_data):
        logger.error("Step 2 failed: Daily load data filling failed.")
        return
    
    logger.info("Step 2 completed successfully")
    
    # Step 3: Fill profile instant data
    logger.info("Initiating Step 3: Profile Instant Data Filling")
    if not fill_profile_instant_data(account_data):
        logger.error("Step 3 failed: Profile instant data filling failed.")
        return
    
    logger.info("Step 3 completed successfully")
    
    # Step 4: Trigger prepaid ledger APIs
    logger.info("Initiating Step 4: Prepaid Ledger API Trigger")
    if not trigger_prepaid_ledger(account_data['accountId']):
        logger.error("Step 4 failed: Prepaid ledger API calls failed.")
        return
    
    logger.info("Step 4 completed successfully")
    
    # Step 5: Generate Excel report
    logger.info("Initiating Step 5: Excel Report Generation")
    if not generate_excel_report(account_data):
        logger.error("Step 5 failed: Excel report generation failed.")
        return
    
    logger.info("Step 5 completed successfully")
    
    # Summary
    logger.info("=" * 80)
    logger.info("PE_182 Test Plan - All Steps Completed Successfully!")
    logger.info("=" * 80)
    logger.info(f"Account ID: {account_data['accountId']}")
    logger.info(f"Meter Serial No: {account_data['meterSrno']}")
    logger.info(f"Meter Install Date: {account_data['meterInstalldate']}")
    logger.info(f"Excel Report: Result_File/Report_PE_182.xlsx")
    logger.info("=" * 80)

if __name__ == "__main__":
    main()
